"""
Excel export: normalize extracted data, filter columns by config, append to extracted_data.xlsx.
"""
import os
import re
import pandas as pd
from datetime import datetime
from openpyxl.styles import PatternFill, Alignment

from core.invoice_validation import validate_invoice_line_row, _parse_number as _inv_parse_num
from core.prior_pieces import apply_prior_year_dash, load_prior_piece_set


def _safe_float(value):
    """Convert to float if possible, else None."""
    if value is None or value == "":
        return None
    try:
        if isinstance(value, (int, float)):
            return float(value)
        s = str(value).strip().replace(",", "")
        return float(s) if s else None
    except (ValueError, TypeError):
        return None


def _excel_number(value):
    """
    Value suitable for an Excel number cell: int/float, or None for blank.
    Whole numbers become int; others keep up to 2 decimal places.
    """
    n = _safe_float(value)
    if n is None:
        return None
    if float(n).is_integer():
        return int(n)
    return round(float(n), 2)


def _apply_numeric_cell_types(ws, numeric_headers: set) -> None:
    """Ensure listed header columns store Python numbers (not text)."""
    if ws is None or ws.max_row < 1:
        return
    header = [c.value for c in ws[1]]
    col_idxs = []
    for name in numeric_headers:
        if name in header:
            col_idxs.append(header.index(name) + 1)
    if not col_idxs:
        return
    for row_idx in range(2, ws.max_row + 1):
        for col_idx in col_idxs:
            cell = ws.cell(row=row_idx, column=col_idx)
            num = _excel_number(cell.value)
            if num is None:
                cell.value = None
            else:
                cell.value = num
                cell.number_format = "0.##"


def _safe_filename_part(value: str, fallback: str = "UNKNOWN") -> str:
    """Sanitize a string for use in a filename."""
    s = str(value or "").strip()
    if not s:
        s = fallback
    s = re.sub(r"[^\w\-.]+", "_", s, flags=re.UNICODE)
    s = re.sub(r"_+", "_", s).strip("._-")
    return s or fallback


def _header_first(header: dict, *keys: str) -> str:
    """Return first non-empty header value among keys."""
    if not isinstance(header, dict):
        return ""
    for k in keys:
        v = header.get(k)
        if v is not None and str(v).strip() != "":
            return str(v).strip()
    return ""


def build_challan_excel_filename(data: dict) -> str:
    """
    Name: <ChallanNo>_<CompanyName>.xlsx
    Company is the From side (mill / supplier), not the To / buyer.
    """
    header = data.get("header", {}) or {}
    challan_no = _header_first(header, "challan_number", "challan_no")
    company = _header_first(
        header,
        "company_name",
        "supplier_name",
        "from_company",
        "mill_name",
    )
    return f"{_safe_filename_part(challan_no, 'NO_CHALLAN')}_{_safe_filename_part(company)}.xlsx"


def build_invoice_excel_filename(data: dict) -> str:
    """Name: <InvoiceNo>_<CompanyName>.xlsx"""
    header = data.get("header", {}) or {}
    inv_no = _header_first(header, "invoice_number", "invoice_no")
    company = _header_first(
        header,
        "bill_to",
        "buyer_name",
        "party_name",
        "supplier_name",
        "company_name",
    )
    return f"{_safe_filename_part(inv_no, 'NO_INVOICE')}_{_safe_filename_part(company)}.xlsx"


def _normalize_document_type(doc_type: str) -> str:
    """Normalize to 'invoice' or 'delivery_challan'."""
    if not doc_type:
        return "unknown"
    t = str(doc_type).strip().lower().replace(" ", "_")
    if "invoice" in t or t == "invoice":
        return "invoice"
    if "challan" in t or "delivery_challan" in t:
        return "delivery_challan"
    return doc_type


def _normalize_header_value(header: dict, key: str) -> str:
    """Get header value as string, format total_invoice_value to 2 decimals."""
    v = header.get(key) or ""
    if v is None:
        return ""
    v = str(v).strip()
    if key == "total_invoice_value" and v:
        fv = _safe_float(v)
        if fv is not None:
            return f"{fv:.2f}"
    return v


def _normalize_item_row(item: dict, file_name: str) -> dict:
    """
    Build one full item row (invoice + challan columns) with normalized numbers.
    Used for both invoice and challan; unused columns will be filtered later.
    """
    def v(key, default=""):
        x = item.get(key) or default
        return "" if x is None else str(x).strip()

    taxable_val = _safe_float(item.get("taxable_value"))
    tax_amt = _safe_float(item.get("tax_amount"))
    total = _safe_float(item.get("total_value"))
    if taxable_val is not None and tax_amt is not None:
        total = round(taxable_val + tax_amt, 2)
    unit_price = _safe_float(item.get("unit_price"))
    quantity = _safe_float(item.get("quantity"))
    if quantity is None:
        quantity = item.get("quantity") or item.get("dispatch_mtr") or ""

    row = {
        "file_name": file_name,
        "item_description": v("item_description") or v("fabric_name"),
        "hsn_code": v("hsn_code"),
        "quantity": f"{quantity:.2f}" if isinstance(quantity, (int, float)) else quantity,
        "uom": v("uom"),
        "unit_price": f"{unit_price:.2f}" if unit_price is not None else v("unit_price"),
        "discount": v("discount"),
        "taxable_value": f"{taxable_val:.2f}" if taxable_val is not None else v("taxable_value"),
        "gst_rate": v("gst_rate"),
        "tax_amount": f"{tax_amt:.2f}" if tax_amt is not None else v("tax_amount"),
        "total_value": f"{total:.2f}" if total is not None else v("total_value"),
        "fabric_name": v("fabric_name"),
        "fd_number": v("fd_number"),
        "piece_number": v("piece_number") or v("piece_no"),
        "challan_mtr": v("challan_mtr"),
        "dispatch_mtr": v("dispatch_mtr") or v("finished_mtrs") or v("fin_mtrs"),
        "grey_mtrs": v("grey_mtrs"),
        "grey_received_date": v("grey_received_date"),
        "grey_challan_number": v("grey_challan_number"),
        "beam": v("beam"),
        "s_no": item.get("s_no", item.get("sno", item.get("serial_no", ""))),
        "flag": bool(item.get("flag", False)),
        "reason": v("reason"),
    }
    return row


def _build_full_doc_row(header: dict, file_name: str, doc_type: str) -> dict:
    """Build document summary row with all possible columns (filtered by config later)."""
    def h(key):
        return _normalize_header_value(header, key)

    total_inv = h("total_invoice_value")
    return {
        "file_name": file_name,
        "document_type": doc_type,
        "invoice_number": h("invoice_number"),
        "invoice_date": h("invoice_date"),
        "supplier_name": h("supplier_name"),
        "supplier_gstin": h("supplier_gstin"),
        "buyer_name": h("buyer_name"),
        "buyer_gstin": h("buyer_gstin"),
        "buyer_address": h("buyer_address"),
        "state_code": h("state_code"),
        "place_of_supply": h("place_of_supply"),
        "total_taxable_value": h("total_taxable_value"),
        "cgst_amount": h("cgst_amount"),
        "sgst_amount": h("sgst_amount"),
        "igst_amount": h("igst_amount"),
        "total_tax": h("total_tax"),
        "total_invoice_value": total_inv,
        "challan_number": h("challan_number"),
        "challan_date": h("challan_date"),
        "company_name": h("company_name"),
        "party_name": h("party_name"),
        "party_address": h("party_address"),
        "hsn_code": h("hsn_code"),
        "total_beam_total": h("total_beam_total"),
        "total_dispatch_mtr": h("total_dispatch_mtr"),
    }


# Column order for sheets.
DOC_STATIC_COLS = ["file_name", "document_type"]
ITEM_STATIC_COLS = []

# Export headers required by challan import format.
EXPORT_PIECE_COL = "Process PieceNo"
EXPORT_GREY_COL = "Grey Mtr"
EXPORT_MTR_COL = "Finish Mtr"
ITEMS_SHEET_NAME = "Sheet1"

# Piece-number characters that should flag a validation row red.
_PIECE_FLAG_CHARS = set("IJLOQVW")


def _piece_has_flag_chars(piece: str) -> bool:
    """True if piece contains I, J, L, O, Q, V, W, or the substring TP (case-insensitive)."""
    s = str(piece or "").upper()
    if "TP" in s:
        return True
    return any(ch in _PIECE_FLAG_CHARS for ch in s)


def _strip_tp_from_piece(piece: str) -> str:
    """
    Remove TP from piece number for Items export (case-insensitive).
    Also removes bracketed forms like (TP), [TP], {TP} so 223ZC(TP) -> 223ZC.
    """
    s = str(piece or "")
    # Drop whole bracketed TP markers first (avoids leftover empty brackets).
    s = re.sub(r"[\(\[\{]\s*TP\s*[\)\]\}]", "", s, flags=re.IGNORECASE)
    # Remove any remaining TP/Tp/tP/tp substrings.
    out = []
    i = 0
    while i < len(s):
        if i + 1 < len(s) and s[i].upper() == "T" and s[i + 1].upper() == "P":
            i += 2
            continue
        out.append(s[i])
        i += 1
    return "".join(out).strip()


def get_document_columns(config: dict) -> list:
    """Ordered list of document columns to write (file_name, document_type + selected fields)."""
    inv = config.get("invoice_fields") or []
    ch = config.get("challan_fields") or []
    # All possible header keys in display order (no duplicates)
    all_header_keys = [
        "invoice_number", "invoice_date", "supplier_name", "supplier_gstin",
        "buyer_name", "buyer_gstin", "buyer_address", "state_code", "place_of_supply",
        "total_taxable_value", "cgst_amount", "sgst_amount", "igst_amount",
        "total_tax", "total_invoice_value",
        "challan_number", "challan_date", "company_name", "party_name",
        "party_address", "hsn_code", "total_beam_total", "total_dispatch_mtr",
    ]
    selected = [k for k in all_header_keys if k in inv or k in ch]
    return DOC_STATIC_COLS + selected


def get_item_columns(config: dict) -> list:
    """Ordered list of item columns to write in final Excel."""
    ch_t = config.get("challan_table_fields") or []
    # Keep only challan import style columns and map to required export headers.
    all_item_keys = ["piece_number", "grey_mtrs", "dispatch_mtr"]
    selected = [k for k in all_item_keys if k in ch_t]
    if not selected:
        selected = ["piece_number", "grey_mtrs", "dispatch_mtr"]
    out = []
    if "piece_number" in selected:
        out.append(EXPORT_PIECE_COL)
    if "grey_mtrs" in selected:
        out.append(EXPORT_GREY_COL)
    if "dispatch_mtr" in selected:
        out.append(EXPORT_MTR_COL)
    # Always keep Grey Mtr between piece and finish when piece+finish are present.
    if EXPORT_PIECE_COL in out and EXPORT_MTR_COL in out and EXPORT_GREY_COL not in out:
        piece_idx = out.index(EXPORT_PIECE_COL)
        out.insert(piece_idx + 1, EXPORT_GREY_COL)
    return ITEM_STATIC_COLS + out


def _line_item_flag_reason(r: dict) -> tuple[bool, str, object]:
    """
    Return (is_flagged, reason, shrinkage_value) for one line item.
    Same rules as Validation_Report.
    """
    piece = str(r.get("piece_number", "") or "").strip()
    grey = _safe_float(r.get("grey_mtrs"))
    finished = _safe_float(r.get("dispatch_mtr"))
    model_flag = bool(r.get("flag", False))
    reason = str(r.get("reason", "") or "").strip()
    issue_parts = []
    shrinkage = ""

    if _piece_has_flag_chars(piece):
        issue_parts.append("piece_no contains I/J/L/O/Q/V/W or TP")

    if grey is not None and finished is not None:
        if finished >= grey:
            issue_parts.append("finished_mtrs is not smaller than grey_mtrs")
        if grey != 0:
            shrink = ((grey - finished) / grey) * 100.0
            shrinkage = round(shrink, 2)
            if shrink < 2 or shrink > 10:
                issue_parts.append("shrinkage outside 2%-10%")
    else:
        missing = []
        if grey is None:
            missing.append("grey_mtrs")
        if finished is None:
            missing.append("finished_mtrs")
        issue_parts.append(f"missing numeric value(s): {', '.join(missing)}")

    final_flag = model_flag or len(issue_parts) > 0
    if not reason and issue_parts:
        reason = "; ".join(issue_parts)
    return final_flag, reason, shrinkage


def _document_s_no(r: dict):
    """Printed S.No. from the challan image (not Sheet1 row index)."""
    raw = r.get("s_no", r.get("sno", r.get("serial_no", "")))
    if raw is None or str(raw).strip() == "":
        return None
    return _excel_number(raw)


def _build_validation_rows(line_items: list, file_name: str) -> list:
    """
    Build validation report rows using extraction flags + deterministic checks.
    Only flagged (wrong) rows are returned.
    S No. is the printed document serial number from the image.
    """
    rows = []
    for r in line_items:
        final_flag, reason, shrinkage = _line_item_flag_reason(r)
        if not final_flag:
            continue
        piece = str(r.get("piece_number", "") or "").strip()
        grey = _safe_float(r.get("grey_mtrs"))
        finished = _safe_float(r.get("dispatch_mtr"))
        rows.append(
            {
                "S No.": _document_s_no(r),
                "file_name": file_name,
                "piece_no": piece,
                "grey_mtrs": _excel_number(grey),
                "finished_mtrs": _excel_number(finished),
                "shrinkage_percent": _excel_number(shrinkage) if shrinkage != "" else None,
                "flag": True,
                "reason": reason,
            }
        )
    return rows


def _apply_red_fill_rows(ws, excel_rows: list) -> None:
    """Fill the given 1-based Excel row numbers with validation red."""
    if ws is None or not excel_rows:
        return
    red_fill = PatternFill(start_color="FDECEC", end_color="FDECEC", fill_type="solid")
    max_col = ws.max_column or 1
    for row_idx in excel_rows:
        for col_idx in range(1, max_col + 1):
            ws.cell(row=row_idx, column=col_idx).fill = red_fill


def _apply_validation_row_colors(ws, start_data_row: int, end_data_row: int) -> None:
    """
    Color rows in Validation_Report red (only wrong/flagged rows are written).
    """
    if end_data_row < start_data_row:
        return
    _apply_red_fill_rows(ws, list(range(start_data_row, end_data_row + 1)))


def _sheet1_rows_for_flagged(line_items: list, data_start_row: int) -> list:
    """
    Excel row numbers on Sheet1 that correspond to flagged line items.
    data_start_row: 1-based row of the first item in this write batch.
    """
    rows = []
    for i, r in enumerate(line_items):
        flagged, _, _ = _line_item_flag_reason(r)
        if flagged:
            rows.append(data_start_row + i)
    return rows


def _apply_left_alignment(ws) -> None:
    """Force left alignment for all cells in the worksheet."""
    alignment = Alignment(horizontal="left", vertical="top")
    for row_idx in range(1, ws.max_row + 1):
        for col_idx in range(1, ws.max_column + 1):
            ws.cell(row=row_idx, column=col_idx).alignment = alignment


VALIDATION_COLUMNS = [
    "S No.",
    "file_name",
    "piece_no",
    "grey_mtrs",
    "finished_mtrs",
    "shrinkage_percent",
    "flag",
    "reason",
]

ITEMS_NUMERIC_HEADERS = {EXPORT_GREY_COL, EXPORT_MTR_COL}
VALIDATION_NUMERIC_HEADERS = {"S No.", "grey_mtrs", "finished_mtrs", "shrinkage_percent"}


def _build_sheet1_and_validation(
    data: dict,
    file_name: str,
    config: dict,
    *,
    totals_note: str = "",
    totals_failed: bool = False,
) -> tuple:
    """
    Build Sheet1 + Validation_Report dataframes and flagged Sheet1 row indexes (1-based data start at 2).
    """
    items = data.get("items", []) or []
    line_items = [_normalize_item_row(it, file_name) for it in items]
    item_cols = get_item_columns(config)
    prior_set = load_prior_piece_set()
    rows_items = []
    for r in line_items:
        row = {}
        if EXPORT_PIECE_COL in item_cols:
            piece = _strip_tp_from_piece(r.get("piece_number", ""))
            row[EXPORT_PIECE_COL] = apply_prior_year_dash(piece, prior_set)
        if EXPORT_GREY_COL in item_cols:
            row[EXPORT_GREY_COL] = _excel_number(r.get("grey_mtrs", ""))
        if EXPORT_MTR_COL in item_cols:
            row[EXPORT_MTR_COL] = _excel_number(r.get("dispatch_mtr", ""))
        rows_items.append(row)

    validation_rows = _build_validation_rows(line_items, file_name)
    if totals_note and totals_failed:
        validation_rows.append(
            {
                "S No.": None,
                "file_name": file_name,
                "piece_no": "",
                "grey_mtrs": None,
                "finished_mtrs": None,
                "shrinkage_percent": None,
                "flag": True,
                "reason": totals_note,
            }
        )

    df_items = pd.DataFrame(rows_items)
    df_validation = pd.DataFrame(validation_rows, columns=VALIDATION_COLUMNS)
    return df_items, df_validation, line_items


def rewrite_challan_excel(
    output_file: str,
    data: dict,
    file_name: str,
    config: dict,
    *,
    totals_note: str = "",
    totals_failed: bool = False,
) -> bool:
    """
    Replace Sheet1 and Validation_Report entirely (used after totals correction).
    """
    df_items, df_validation, line_items = _build_sheet1_and_validation(
        data,
        file_name,
        config,
        totals_note=totals_note,
        totals_failed=totals_failed,
    )
    with pd.ExcelWriter(output_file, engine="openpyxl", mode="w") as writer:
        df_items.to_excel(writer, sheet_name=ITEMS_SHEET_NAME, index=False)
        df_validation.to_excel(writer, sheet_name="Validation_Report", index=False)
        ws_v = writer.sheets.get("Validation_Report")
        if ws_v is not None and len(df_validation) > 0:
            _apply_validation_row_colors(ws_v, 2, 1 + len(df_validation))
        ws_i = writer.sheets.get(ITEMS_SHEET_NAME)
        if ws_i is not None:
            _apply_red_fill_rows(ws_i, _sheet1_rows_for_flagged(line_items, 2))
            _apply_numeric_cell_types(ws_i, ITEMS_NUMERIC_HEADERS)
            _apply_left_alignment(ws_i)
        if ws_v is not None:
            _apply_numeric_cell_types(ws_v, VALIDATION_NUMERIC_HEADERS)
            _apply_left_alignment(ws_v)
    return True


def write_to_excel(
    data: dict,
    file_name: str,
    output_dir: str,
    config: dict,
    output_file: str = None,
) -> bool:
    """
    Write one document's extracted data to an Excel file.
    Creates the file if missing; appends Sheet1 / Validation_Report if it already exists
    (e.g. continuation page of the same challan).
    Validation_Report includes only wrong (flagged) rows.
    """
    output_file = output_file or os.path.join(output_dir, "extracted_data.xlsx")
    df_items, df_validation, line_items = _build_sheet1_and_validation(
        data, file_name, config
    )

    try:
        if not os.path.exists(output_file):
            with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
                df_items.to_excel(writer, sheet_name=ITEMS_SHEET_NAME, index=False)
                df_validation.to_excel(writer, sheet_name="Validation_Report", index=False)
                ws_v = writer.sheets.get("Validation_Report")
                if ws_v is not None and len(df_validation) > 0:
                    _apply_validation_row_colors(ws_v, 2, 1 + len(df_validation))
                ws_i = writer.sheets.get(ITEMS_SHEET_NAME)
                if ws_i is not None:
                    # Header is row 1; first item is row 2.
                    _apply_red_fill_rows(ws_i, _sheet1_rows_for_flagged(line_items, 2))
                    _apply_numeric_cell_types(ws_i, ITEMS_NUMERIC_HEADERS)
                    _apply_left_alignment(ws_i)
                if ws_v is not None:
                    _apply_numeric_cell_types(ws_v, VALIDATION_NUMERIC_HEADERS)
                    _apply_left_alignment(ws_v)
        else:
            with pd.ExcelWriter(output_file, engine="openpyxl", mode="a", if_sheet_exists="overlay") as writer:
                items_sheet = (
                    ITEMS_SHEET_NAME
                    if ITEMS_SHEET_NAME in writer.sheets
                    else ("Items" if "Items" in writer.sheets else ITEMS_SHEET_NAME)
                )
                if items_sheet not in writer.sheets:
                    df_items.to_excel(writer, sheet_name=ITEMS_SHEET_NAME, index=False)
                    ws_i = writer.sheets.get(ITEMS_SHEET_NAME)
                    sheet1_data_start = 2
                else:
                    startrow = writer.sheets[items_sheet].max_row
                    df_items.to_excel(
                        writer,
                        sheet_name=items_sheet,
                        startrow=startrow,
                        index=False,
                        header=False,
                    )
                    ws_i = writer.sheets.get(items_sheet)
                    # pandas writes first data row at startrow + 1
                    sheet1_data_start = startrow + 1
                if ws_i is not None:
                    _apply_red_fill_rows(
                        ws_i, _sheet1_rows_for_flagged(line_items, sheet1_data_start)
                    )
                    _apply_numeric_cell_types(ws_i, ITEMS_NUMERIC_HEADERS)
                    _apply_left_alignment(ws_i)
                if "Validation_Report" not in writer.sheets:
                    df_validation.to_excel(writer, sheet_name="Validation_Report", index=False)
                    ws_v = writer.sheets.get("Validation_Report")
                    if ws_v is not None and len(df_validation) > 0:
                        _apply_validation_row_colors(ws_v, 2, 1 + len(df_validation))
                else:
                    startrow = writer.sheets["Validation_Report"].max_row
                    if len(df_validation) > 0:
                        df_validation.to_excel(
                            writer,
                            sheet_name="Validation_Report",
                            startrow=startrow,
                            index=False,
                            header=False,
                        )
                        ws_v = writer.sheets.get("Validation_Report")
                        if ws_v is not None:
                            _apply_validation_row_colors(ws_v, startrow + 1, startrow + len(df_validation))
                    else:
                        ws_v = writer.sheets.get("Validation_Report")
                if ws_v is not None:
                    _apply_numeric_cell_types(ws_v, VALIDATION_NUMERIC_HEADERS)
                    _apply_left_alignment(ws_v)
        return True
    except Exception:
        raise


def _invoice_h(header: dict, *keys: str) -> str:
    """Get first non-empty header value from alias keys."""
    if not isinstance(header, dict):
        return ""
    for k in keys:
        v = header.get(k)
        if v is not None and str(v).strip() != "":
            return str(v).strip()
    return ""


def _normalize_invoice_item(item: dict) -> dict:
    """Normalize invoice table row to required output columns."""
    def pick(*keys):
        for k in keys:
            v = item.get(k)
            if v is not None and str(v).strip() != "":
                return str(v).strip()
        return ""
    quality = pick("quality", "item_description", "fabric_name")
    fin = pick("finished_mtrs", "fin_mtrs", "dispatch_mtr")
    rate = pick("rate", "unit_price")
    amount = pick("amount", "Amount", "line_amount", "total_value")
    model_flag = bool(item.get("flag", False))
    model_reason = str(item.get("reason", "") or "").strip()
    final_flag, reason = validate_invoice_line_row(
        quality, fin, rate, amount if amount else None, model_flag=model_flag, model_reason=model_reason
    )
    return {
        "Quality": quality,
        "Fin. Mtrs": fin,
        "Rate": rate,
        "Amount": amount,
        "flag": final_flag,
        "reason": reason,
    }


def write_invoice_to_excel(
    data: dict,
    file_name: str,
    output_file: str,
) -> bool:
    """
    Append one invoice extraction to invoice Excel file.
    Keeps invoice data separate from delivery challan output.
    """
    header = data.get("header", {}) or {}
    items = data.get("items", []) or []

    doc_row = {
        "file_name": file_name,
        "document_type": "invoice",
        "supplier_name": _invoice_h(header, "supplier_name", "company_name"),
        "supplier_gstin": _invoice_h(header, "supplier_gstin", "gstin"),
        "bill_to": _invoice_h(header, "bill_to", "buyer_name", "party_name"),
        "bill_to_gstin": _invoice_h(header, "bill_to_gstin", "buyer_gstin"),
        "invoice_number": _invoice_h(header, "invoice_number"),
        "invoice_date": _invoice_h(header, "invoice_date", "dated"),
        "challan_number": _invoice_h(header, "challan_number"),
        "ewb_no": _invoice_h(header, "ewb_no"),
        "ack_no": _invoice_h(header, "ack_no"),
        "irn": _invoice_h(header, "irn"),
        "state_code": _invoice_h(header, "state_code"),
    }
    item_rows = [_normalize_invoice_item(it) for it in items if isinstance(it, dict)]
    df_doc = pd.DataFrame([doc_row])
    df_items = pd.DataFrame(item_rows, columns=["Quality", "Fin. Mtrs", "Rate", "Amount", "flag", "reason"])

    # Validation sheet (one row per line item)
    val_rows = []
    for idx, row in enumerate(item_rows, start=1):
        fin = _inv_parse_num(row.get("Fin. Mtrs"))
        r = _inv_parse_num(row.get("Rate"))
        amt = _inv_parse_num(row.get("Amount"))
        expected = ""
        if fin is not None and r is not None:
            expected = round(fin * r, 2)
        val_rows.append(
            {
                "file_name": file_name,
                "row": idx,
                "Quality": row.get("Quality", ""),
                "Fin. Mtrs": row.get("Fin. Mtrs", ""),
                "Rate": row.get("Rate", ""),
                "Amount": row.get("Amount", ""),
                "expected_amount": expected,
                "flag": row.get("flag", False),
                "reason": row.get("reason", ""),
            }
        )
    df_val = pd.DataFrame(val_rows)

    try:
        if not os.path.exists(output_file):
            with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
                df_doc.to_excel(writer, sheet_name="Invoice_Documents", index=False)
                df_items.to_excel(writer, sheet_name="Invoice_Items", index=False)
                df_val.to_excel(writer, sheet_name="Invoice_Validation", index=False)
                ws_v = writer.sheets.get("Invoice_Validation")
                if ws_v is not None and len(df_val) > 0:
                    _apply_validation_row_colors(ws_v, 2, 1 + len(df_val))
                ws_d = writer.sheets.get("Invoice_Documents")
                if ws_d is not None:
                    _apply_left_alignment(ws_d)
                ws_i = writer.sheets.get("Invoice_Items")
                if ws_i is not None:
                    _apply_left_alignment(ws_i)
                if ws_v is not None:
                    _apply_left_alignment(ws_v)
        else:
            with pd.ExcelWriter(output_file, engine="openpyxl", mode="a", if_sheet_exists="overlay") as writer:
                if "Invoice_Documents" not in writer.sheets:
                    df_doc.to_excel(writer, sheet_name="Invoice_Documents", index=False)
                else:
                    startrow = writer.sheets["Invoice_Documents"].max_row
                    df_doc.to_excel(writer, sheet_name="Invoice_Documents", startrow=startrow, index=False, header=False)

                if "Invoice_Items" not in writer.sheets:
                    df_items.to_excel(writer, sheet_name="Invoice_Items", index=False)
                else:
                    startrow = writer.sheets["Invoice_Items"].max_row
                    df_items.to_excel(writer, sheet_name="Invoice_Items", startrow=startrow, index=False, header=False)

                ws_d = writer.sheets.get("Invoice_Documents")
                if ws_d is not None:
                    _apply_left_alignment(ws_d)
                ws_i = writer.sheets.get("Invoice_Items")
                if ws_i is not None:
                    _apply_left_alignment(ws_i)

                if "Invoice_Validation" not in writer.sheets:
                    df_val.to_excel(writer, sheet_name="Invoice_Validation", index=False)
                    ws_v = writer.sheets.get("Invoice_Validation")
                    if ws_v is not None and len(df_val) > 0:
                        _apply_validation_row_colors(ws_v, 2, 1 + len(df_val))
                else:
                    startrow = writer.sheets["Invoice_Validation"].max_row
                    df_val.to_excel(writer, sheet_name="Invoice_Validation", startrow=startrow, index=False, header=False)
                    ws_v = writer.sheets.get("Invoice_Validation")
                    if ws_v is not None and len(df_val) > 0:
                        _apply_validation_row_colors(ws_v, startrow + 1, startrow + len(df_val))
                if ws_v is not None:
                    _apply_left_alignment(ws_v)
        return True
    except Exception:
        raise
