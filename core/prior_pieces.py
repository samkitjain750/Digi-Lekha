"""
Prior-year (due) piece numbers: load, save, import from Excel, match for Sheet1 dash prefix.
"""
from __future__ import annotations

import json
import os
import re
from datetime import datetime

from . import paths as _paths

PRIOR_PIECES_JSON = "prior_year_pieces.json"
PRIOR_PIECES_XLSX = "prior_year_pieces.xlsx"

_PIECE_HEADER_ALIASES = {
    "process pieceno",
    "process piece no",
    "piece no",
    "piece no.",
    "pieceno",
    "piece_number",
    "piece number",
}


def _strip_tp_local(piece: str) -> str:
    s = str(piece or "")
    s = re.sub(r"[\(\[\{]\s*TP\s*[\)\]\}]", "", s, flags=re.IGNORECASE)
    out = []
    i = 0
    while i < len(s):
        if i + 1 < len(s) and s[i].upper() == "T" and s[i + 1].upper() == "P":
            i += 2
            continue
        out.append(s[i])
        i += 1
    return "".join(out).strip()


def _normalize_header(value) -> str:
    s = str(value or "").strip().lower()
    s = s.replace("_", " ")
    s = re.sub(r"\s+", " ", s)
    return s


def get_prior_pieces_json_path() -> str:
    return os.path.join(_paths.get_config_dir(writable=True), PRIOR_PIECES_JSON)


def get_prior_pieces_xlsx_path() -> str:
    return os.path.join(_paths.get_config_dir(writable=True), PRIOR_PIECES_XLSX)


def normalize_piece_key(piece: str) -> str:
    """Canonical key for matching (no leading dash, uppercased, TP stripped)."""
    s = _strip_tp_local(str(piece or "").strip())
    s = s.lstrip("-").strip()
    s = re.sub(r"-\d+$", "", s)
    return s.upper()


def _looks_like_piece(value: str) -> bool:
    s = str(value or "").strip()
    if not s:
        return False
    low = s.lower()
    if low in _PIECE_HEADER_ALIASES or "quality" in low or "total" in low:
        return False
    try:
        float(s.replace(",", ""))
        return False
    except ValueError:
        pass
    return bool(re.search(r"[A-Za-z]", s))


def load_prior_piece_set() -> set[str]:
    """Return set of normalized piece keys from saved list (empty if none)."""
    path = get_prior_pieces_json_path()
    if not os.path.isfile(path):
        seeded = _try_seed_from_default_xlsx()
        if seeded:
            return seeded
        return set()
    try:
        with open(path, "r", encoding="utf-8") as f:
            data = json.load(f)
        pieces = data.get("pieces") or []
        return {normalize_piece_key(p) for p in pieces if normalize_piece_key(p)}
    except Exception:
        return set()


def load_prior_pieces_meta() -> dict:
    """Metadata for UI: count, updated_at, source_file."""
    path = get_prior_pieces_json_path()
    if not os.path.isfile(path):
        return {"count": 0, "updated_at": "", "source_file": ""}
    try:
        with open(path, "r", encoding="utf-8") as f:
            data = json.load(f)
        pieces = data.get("pieces") or []
        return {
            "count": len(pieces),
            "updated_at": data.get("updated_at", ""),
            "source_file": data.get("source_file", ""),
        }
    except Exception:
        return {"count": 0, "updated_at": "", "source_file": ""}


def save_prior_pieces(pieces: list[str], source_file: str = "") -> dict:
    """Persist piece list as JSON + Excel copy. Returns meta."""
    cleaned = []
    seen = set()
    for p in pieces:
        key = normalize_piece_key(p)
        if not key or key in seen:
            continue
        seen.add(key)
        display = _strip_tp_local(str(p).strip()).lstrip("-").strip()
        display = re.sub(r"-\d+$", "", display)
        cleaned.append(display)

    meta = {
        "updated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "source_file": os.path.basename(source_file) if source_file else "",
        "pieces": cleaned,
    }
    json_path = get_prior_pieces_json_path()
    os.makedirs(os.path.dirname(json_path), exist_ok=True)
    with open(json_path, "w", encoding="utf-8") as f:
        json.dump(meta, f, indent=2, ensure_ascii=False)

    try:
        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        ws["A1"] = "Process PieceNo"
        for i, p in enumerate(cleaned, start=2):
            ws.cell(row=i, column=1, value=p)
        wb.save(get_prior_pieces_xlsx_path())
    except Exception:
        pass

    return {
        "count": len(cleaned),
        "updated_at": meta["updated_at"],
        "source_file": meta["source_file"],
    }


def _try_seed_from_default_xlsx() -> set[str]:
    """If no saved list yet, import project-root piece_numbers.xlsx when available."""
    candidates = [
        os.path.join(_paths.get_writable_base(), "piece_numbers.xlsx"),
        os.path.join(_paths.get_resource_base(), "piece_numbers.xlsx"),
    ]
    for path in candidates:
        if os.path.isfile(path):
            try:
                pieces, _ = extract_pieces_from_excel(path)
                save_prior_pieces(pieces, source_file=path)
                return {normalize_piece_key(p) for p in pieces}
            except Exception:
                continue
    return set()


def extract_pieces_from_excel(file_path: str) -> tuple[list[str], str]:
    """
    Parse an uploaded Excel of prior-year piece numbers.
    Returns (pieces, used_column_name).
    Raises ValueError with a user-facing message on invalid format.
    """
    import pandas as pd

    if not file_path or not os.path.isfile(file_path):
        raise ValueError("File not found.")

    ext = file_path.lower().rsplit(".", 1)[-1]
    if ext not in ("xlsx", "xls"):
        raise ValueError("Please upload an Excel file (.xlsx or .xls).")

    try:
        if ext == "xls":
            df = pd.read_excel(file_path, header=None, engine="xlrd")
        else:
            df = pd.read_excel(file_path, header=None, engine="openpyxl")
    except Exception as e:
        raise ValueError(
            f"Could not read Excel file. Make sure it is a valid .xlsx/.xls.\nDetails: {e}"
        ) from e

    if df is None or df.empty:
        raise ValueError("The Excel sheet is empty. Add piece numbers under column A.")

    df = df.dropna(axis=1, how="all")
    if df.shape[1] == 0:
        raise ValueError("No columns found. Use one column: Process PieceNo.")

    piece_col = None
    header_row = None
    used_name = "Process PieceNo"
    max_scan = min(5, len(df))
    for r in range(max_scan):
        for c in range(df.shape[1]):
            cell = df.iat[r, c]
            if pd.isna(cell):
                continue
            if _normalize_header(cell) in _PIECE_HEADER_ALIASES:
                piece_col = c
                header_row = r
                used_name = str(cell).strip()
                break
        if piece_col is not None:
            break

    if piece_col is None:
        if df.shape[1] == 1:
            piece_col = 0
            first = df.iat[0, 0]
            if not pd.isna(first) and _normalize_header(first) in _PIECE_HEADER_ALIASES:
                header_row = 0
                used_name = str(first).strip()
            elif not pd.isna(first) and not _looks_like_piece(str(first)):
                header_row = 0
                used_name = "Process PieceNo"
            else:
                header_row = None
        else:
            raise ValueError(
                "Wrong Excel format.\n\n"
                "Expected: one column of piece numbers.\n"
                "Header (optional): Process PieceNo  or  Piece No\n\n"
                "Your file has multiple columns and no Piece No header was found.\n"
                "Click ? to see the correct format."
            )

    start = (header_row + 1) if header_row is not None else 0
    pieces = []
    for r in range(start, len(df)):
        raw = df.iat[r, piece_col]
        if pd.isna(raw):
            continue
        s = str(raw).strip()
        if not _looks_like_piece(s):
            continue
        pieces.append(s)

    if not pieces:
        raise ValueError(
            "No valid piece numbers found.\n\n"
            "Expected rows like: 1431ZN, 30707J\n"
            "(Must contain letters, not only numbers.)\n"
            "Click ? to see the correct format."
        )

    return pieces, used_name


def apply_prior_year_dash(piece: str, prior_set: set[str] | None = None) -> str:
    """
    If piece matches a prior-year due piece, ensure it starts with '-'.
    Does not remove an existing leading '-'.
    """
    if prior_set is None:
        prior_set = load_prior_piece_set()
    cleaned = _strip_tp_local(str(piece or "").strip())
    if not cleaned:
        return cleaned
    key = normalize_piece_key(cleaned)
    if key and key in prior_set:
        if not cleaned.startswith("-"):
            return "-" + cleaned
    return cleaned


def format_help_text() -> str:
    return (
        "Prior-year due pieces — Excel format\n"
        "====================================\n\n"
        "Upload a simple Excel file (.xlsx or .xls) with piece numbers only.\n\n"
        "Required layout:\n"
        "  • One column (column A)\n"
        "  • Header (recommended): Process PieceNo\n"
        "    Also accepted: Piece No, PieceNo\n"
        "  • Each next row: one piece number\n\n"
        "Example:\n"
        "  Process PieceNo\n"
        "  1431ZN\n"
        "  30707J\n"
        "  4486ZS\n\n"
        "Notes:\n"
        "  • Do not include Grey Mtr / Finish Mtr columns\n"
        "  • Piece numbers must include letters (e.g. 30707J)\n"
        "  • This list is saved and reused on future extractions\n"
        "  • Matching pieces get a '-' prefix on Sheet1\n"
    )
